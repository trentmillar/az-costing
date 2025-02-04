from azure.identity import DefaultAzureCredential
from azure.mgmt.subscription import SubscriptionClient
from azure.mgmt.consumption import ConsumptionManagementClient
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import pandas as pd
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill
from time import sleep
from azure.core.exceptions import HttpResponseError
import random
import requests
from plotly import graph_objects as go
from plotly.subplots import make_subplots
import io
from openpyxl.drawing.image import Image
from scipy import stats
import numpy as np
from azure.mgmt.resource import ResourceManagementClient
from azure.mgmt.advisor import AdvisorManagementClient

def get_subscriptions():
    """Get all subscriptions in the tenant"""
    credential = DefaultAzureCredential()
    subscription_client = SubscriptionClient(credential)
    
    subscriptions = []
    for sub in subscription_client.subscriptions.list():
        if sub.display_name.lower().startswith('sub'):
            subscriptions.append({
                'id': sub.subscription_id,
                'name': sub.display_name
            })
    return subscriptions

def get_costs_by_resource_type(subscription_id, start_date, end_date, max_retries=5, window_size=7):
    """Get costs grouped by resource type for a subscription in a date range"""
    credential = DefaultAzureCredential()
    token = credential.get_token("https://management.azure.com/.default")
    auth_header = f"Bearer {token.token}"
    
    headers = {
        "Authorization": auth_header,
        "Content-Type": "application/json"
    }   
    
    def fetch_costs_for_window(start, end, attempt=0):
        """Helper function to fetch costs for a specific window"""
        try:
            print(f"Getting costs for window: {start} to {end} on {subscription_id}")
            
            url = f"https://management.azure.com/subscriptions/{subscription_id}/providers/Microsoft.Consumption/usageDetails"
            params = {
                "api-version": "2019-11-01",
                "startDate": start,
                "endDate": end
            }
            
            session = requests.Session()
            adapter = requests.adapters.HTTPAdapter(
                max_retries=requests.urllib3.Retry(
                    total=5,
                    backoff_factor=1,
                    status_forcelist=[429, 500, 502, 503, 504],
                    allowed_methods=["HEAD", "GET", "OPTIONS"]
                ),
                pool_connections=10,
                pool_maxsize=10
            )
            session.mount('https://', adapter)
            
            response = session.get(url, headers=headers, params=params, timeout=90)
            response.raise_for_status()
            
            return response.json()
            
        except Exception as e:
            # Calculate the current window size in days
            current_window = (datetime.strptime(end, '%Y-%m-%d') - datetime.strptime(start, '%Y-%m-%d')).days
            
            # If window is more than 1 day, try splitting
            if current_window > 1:
                print(f"Error with {current_window} day window, splitting into smaller windows...")
                print(f"Error details: {str(e)}")
                
                # Calculate mid point
                mid_date = datetime.strptime(start, '%Y-%m-%d') + timedelta(days=current_window // 2)
                mid_str = mid_date.strftime('%Y-%m-%d')
                
                print(f"Splitting into: {start} to {mid_str} and {mid_str} to {end}")
                
                # Recursively fetch both halves
                first_half = fetch_costs_for_window(start, mid_str)
                second_half = fetch_costs_for_window(mid_str, end)
                
                # Combine the results
                combined = {'value': []}
                if first_half and 'value' in first_half:
                    combined['value'].extend(first_half['value'])
                if second_half and 'value' in second_half:
                    combined['value'].extend(second_half['value'])
                return combined
            
            # If window is 1 day and still failing, try regular retry logic
            elif attempt < max_retries - 1:
                wait_time = min(30 * (attempt + 1), 300)
                print(f"Request failed for 1-day window: {str(e)}. Retrying in {wait_time} seconds... ({attempt + 1}/{max_retries})")
                sleep(wait_time)
                return fetch_costs_for_window(start, end, attempt + 1)
            else:
                raise
    
    # Initial fetch with the full window
    usage = fetch_costs_for_window(start_date, end_date)
    
    # Aggregate costs by resource type
    costs_by_type = {}
    total_cost = 0
    for item in usage.get('value', []):
        resource_type = item.get('properties', {}).get('serviceFamily') or "Unknown"
        cost = float(item.get('properties', {}).get('costInBillingCurrency', 0))
        costs_by_type[resource_type] = costs_by_type.get(resource_type, 0) + cost
        total_cost += cost
    
    print(f"Total cost for period: ${total_cost:,.2f}")
    return costs_by_type

def get_subscription_tags(subscription_id, credential):
    """Get all unique tags used in a subscription"""
    resource_client = ResourceManagementClient(credential, subscription_id)
    
    tags = {}
    for resource in resource_client.resources.list():
        if resource.tags:
            for key, value in resource.tags.items():
                if key not in tags:
                    tags[key] = set()
                tags[key].add(value)
    
    return tags

def create_cost_visualizations(wb, monthly_costs):
    """Create visualization sheet with cost analysis charts"""
    # Create a new sheet for visualizations
    viz_sheet = wb.create_sheet("Visualizations", 0)  # Put it first
    
    # Prepare data for plotting
    periods = sorted(monthly_costs.keys())
    period_labels = [f"{calendar.month_name[m]} {y}" for y, m in periods]
    
    # 1. Total costs per month (line chart)
    total_costs_by_month = []
    for period in periods:
        total = sum(
            service_costs  # Changed: service_costs is already a float
            for sub_costs in monthly_costs[period].values()
            for service_costs in sub_costs.values()
        )
        total_costs_by_month.append(total)
    
    # 2. Costs by subscription (stacked bar chart)
    subscription_names = {sub['id']: sub['name'] for sub in get_subscriptions()}
    sub_costs_by_month = {sub_id: [] for sub_id in subscription_names}
    
    # Calculate total cost for each subscription across all months
    sub_total_costs = {}
    for sub_id in subscription_names:
        total_cost = sum(
            monthly_costs[period].get(sub_id, {}).get(service_family, 0)
            for period in periods
            for service_family in monthly_costs[period].get(sub_id, {})
        )
        sub_total_costs[sub_id] = total_cost
    
    # Get top 50 subscriptions by cost
    top_subs = dict(sorted(sub_total_costs.items(), key=lambda x: x[1], reverse=True)[:25])
    
    # Calculate costs by month for top subscriptions and group others
    sub_costs_by_month = {sub_id: [] for sub_id in top_subs}
    other_costs_by_month = []
    
    for period in periods:
        # Calculate costs for top subscriptions
        for sub_id in top_subs:
            total = sum(
                service_costs
                for service_family, service_costs in monthly_costs[period].get(sub_id, {}).items()
            )
            sub_costs_by_month[sub_id].append(total)
        
        # Calculate combined costs for other subscriptions
        other_total = sum(
            service_costs
            for sub_id, sub_data in monthly_costs[period].items()
            for service_family, service_costs in sub_data.items()
            if sub_id not in top_subs
        )
        other_costs_by_month.append(other_total)
    
    # Add "Others" to the visualization
    sub_costs_by_month['Others'] = other_costs_by_month
    
    # Create subplot figure first
    fig = make_subplots(
        rows=2, cols=2,
        specs=[[{"type": "scatter"}, {"type": "bar"}],
               [{"type": "pie"}, {"type": "scatter"}]],
        subplot_titles=(
            "Monthly Total Costs Trend",
            "Monthly Costs by Subscription",
            f"Service Distribution ({period_labels[-1]})",
            "Cost Growth Rate"
        )
    )
    
    # 1. Line chart - Total costs trend
    fig.add_trace(
        go.Scatter(
            x=period_labels,
            y=total_costs_by_month,
            mode='lines+markers',
            name='Total Cost'
        ),
        row=1, col=1
    )
    
    # 2. Stacked bar chart - Costs by subscription (top 10 + Others)
    for sub_id, costs in sub_costs_by_month.items():
        name = subscription_names.get(sub_id, 'Others')  # Use 'Others' for the aggregated data
        fig.add_trace(
            go.Bar(
                x=period_labels,
                y=costs,
                name=name if len(name) <= 30 else name[:27] + '...'  # Truncate long names
            ),
            row=1, col=2
        )
    
    # 3. Service distribution for latest month (pie chart)
    latest_period = periods[-1]
    service_totals = {}
    for sub_costs in monthly_costs[latest_period].values():
        for service_family, cost in sub_costs.items():
            service_totals[service_family] = service_totals.get(service_family, 0) + cost
    
    # Sort services by cost and get top 10
    top_services = dict(sorted(service_totals.items(), key=lambda x: x[1], reverse=True)[:10])
    other_services = sum(cost for service, cost in service_totals.items() if service not in top_services)
    
    # Add "Others" category
    if other_services > 0:
        top_services['Others'] = other_services
    
    fig.add_trace(
        go.Pie(
            labels=list(top_services.keys()),
            values=list(top_services.values()),
            hole=0.3,
            textinfo='label+percent',
            showlegend=False
        ),
        row=2, col=1
    )
    
    # 4. Growth rate chart
    growth_rates = [
        ((total_costs_by_month[i] - total_costs_by_month[i-1]) / total_costs_by_month[i-1] * 100)
        if total_costs_by_month[i-1] != 0 else 0
        for i in range(1, len(total_costs_by_month))
    ]
    
    fig.add_trace(
        go.Scatter(
            x=period_labels[1:],
            y=growth_rates,
            mode='lines+markers',
            name='Growth Rate (%)'
        ),
        row=2, col=2
    )
    
    # Update layout
    fig.update_layout(
        height=800,
        showlegend=True,
        title_text="Azure Cost Analysis"
    )
    
    # Save plot as image and embed in Excel
    img_bytes = fig.to_image(format="png", width=1200, height=800)
    img = Image(io.BytesIO(img_bytes))
    
    # Add image to sheet
    viz_sheet.add_image(img, 'A1')
    
    # Adjust column widths and row heights to fit the image
    viz_sheet.column_dimensions['A'].width = 120
    viz_sheet.row_dimensions[1].height = 600

def create_anomaly_analysis(wb, monthly_costs):
    """Create anomaly analysis sheet using statistical methods"""
    print("\nStarting anomaly analysis...")
    
    # Create a new sheet for anomaly detection
    anomaly_sheet = wb.create_sheet("Anomaly Analysis", 1)
    
    # Set up headers
    headers = ['Subscription', 'Period', 'Cost', 'Z-Score', 'Is Anomaly', 'Previous Period Cost', 'Cost Change', '% Change']
    for col, header in enumerate(headers, 1):
        cell = anomaly_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Get subscription names
    subscription_names = {sub['id']: sub['name'] for sub in get_subscriptions()}
    periods = sorted(monthly_costs.keys())
    period_labels = [f"{calendar.month_name[m]} {y}" for y, m in periods]
    
    print(f"Analyzing {len(subscription_names)} subscriptions over {len(periods)} periods")
    
    row = 2  # Start after headers
    
    # Analyze each subscription
    for sub_id in monthly_costs[periods[0]].keys():  # Use first period to get sub list
        # Get cost history for this subscription
        costs = []
        for period in periods:
            # Sum all service costs for this subscription in this period
            total = sum(
                cost
                for service_family, cost in monthly_costs[period].get(sub_id, {}).items()
            )
            costs.append(total)
        
        print(f"\nSubscription: {subscription_names.get(sub_id, sub_id)}")
        print(f"Cost history: {costs}")
        
        # Convert to numpy array
        costs_array = np.array(costs)
        
        # Only perform analysis if we have enough data points and non-zero variance
        if len(costs_array) >= 3 and np.std(costs_array) > 0:
            # Calculate z-scores
            z_scores = stats.zscore(costs_array)
            print(f"Z-scores: {z_scores}")
            
            # Identify anomalies (beyond 2 standard deviations)
            for i, (cost, z_score) in enumerate(zip(costs_array, z_scores)):
                is_anomaly = abs(z_score) > 2
                
                if is_anomaly:
                    print(f"Found anomaly: Period={period_labels[i]}, Cost=${cost:,.2f}, Z-score={z_score:.2f}")
                
                # Only add to sheet if it's an anomaly
                if is_anomaly:
                    # Calculate cost change from previous period
                    prev_cost = costs_array[i-1] if i > 0 else 0
                    cost_change = cost - prev_cost
                    pct_change = (cost_change / prev_cost * 100) if prev_cost > 0 else float('inf')
                    
                    # Add to sheet
                    anomaly_sheet.cell(row=row, column=1).value = subscription_names.get(sub_id, sub_id)
                    anomaly_sheet.cell(row=row, column=2).value = period_labels[i]
                    anomaly_sheet.cell(row=row, column=3).value = cost
                    anomaly_sheet.cell(row=row, column=3).number_format = '$#,##0.00'
                    anomaly_sheet.cell(row=row, column=4).value = round(z_score, 2)
                    anomaly_sheet.cell(row=row, column=5).value = "High" if z_score > 2 else "Low"
                    anomaly_sheet.cell(row=row, column=6).value = prev_cost
                    anomaly_sheet.cell(row=row, column=6).number_format = '$#,##0.00'
                    anomaly_sheet.cell(row=row, column=7).value = cost_change
                    anomaly_sheet.cell(row=row, column=7).number_format = '$#,##0.00'
                    anomaly_sheet.cell(row=row, column=8).value = f"{pct_change:,.1f}%"
                    
                    # Add conditional formatting
                    if z_score > 2:
                        anomaly_sheet.cell(row=row, column=5).fill = PatternFill(
                            start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
                        )
                    else:
                        anomaly_sheet.cell(row=row, column=5).fill = PatternFill(
                            start_color="90EE90", end_color="90EE90", fill_type="solid"
                        )
                    
                    row += 1

def create_tag_analysis(wb, monthly_costs, credential):
    """Create sheet analyzing costs by tags"""
    print("\nStarting tag analysis...")
    
    # Create a new sheet for tag analysis
    tag_sheet = wb.create_sheet("Tag Analysis", 2)  # After Anomaly Analysis
    
    # Get all unique tags across subscriptions
    all_tags = {}
    for sub_id in monthly_costs[list(monthly_costs.keys())[0]].keys():
        print(f"\nGetting tags for subscription: {sub_id}")
        sub_tags = get_subscription_tags(sub_id, credential)
        for tag_key, tag_values in sub_tags.items():
            if tag_key not in all_tags:
                all_tags[tag_key] = set()
            all_tags[tag_key].update(tag_values)
    
    # Convert to DataFrame for easier analysis
    data = []
    periods = sorted(monthly_costs.keys())
    
    for period in periods:
        year, month = period
        for sub_id, services in monthly_costs[period].items():
            sub_tags = get_subscription_tags(sub_id, credential)
            
            # Calculate total cost for this subscription in this period
            total_cost = sum(
                service_cost
                for service_family, service_cost in services.items()
            )
            
            if sub_tags:
                # If there are tags, distribute cost evenly across tag combinations
                num_tag_combinations = sum(len(values) for values in sub_tags.values())
                cost_per_tag = total_cost / num_tag_combinations if num_tag_combinations > 0 else total_cost
                
                # Add a row for each tag value
                for tag_key, tag_values in sub_tags.items():
                    for tag_value in tag_values:
                        data.append({
                            'Period': f"{calendar.month_name[month]} {year}",
                            'Subscription': sub_id,
                            'Tag Key': tag_key,
                            'Tag Value': tag_value,
                            'Cost': cost_per_tag
                        })
            else:
                # If no tags, mark as untagged
                data.append({
                    'Period': f"{calendar.month_name[month]} {year}",
                    'Subscription': sub_id,
                    'Tag Key': 'Untagged',
                    'Tag Value': 'Untagged',
                    'Cost': total_cost
                })
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create pivot tables
    # 1. Costs by Tag Key
    pivot_tag_keys = pd.pivot_table(
        df,
        values='Cost',
        index='Tag Key',
        columns='Period',
        aggfunc='sum',
        fill_value=0
    )
    
    # 2. Costs by Tag Value for each Tag Key
    tag_value_pivots = {}
    for tag_key in all_tags:
        tag_value_pivots[tag_key] = pd.pivot_table(
            df[df['Tag Key'] == tag_key],
            values='Cost',
            index='Tag Value',
            columns='Period',
            aggfunc='sum',
            fill_value=0
        )
    
    # Write to Excel
    # Header
    tag_sheet['A1'] = 'Tag Analysis'
    tag_sheet['A1'].font = Font(bold=True, size=14)
    
    # Summary by Tag Key
    current_row = 3
    tag_sheet[f'A{current_row}'] = 'Costs by Tag Key'
    tag_sheet[f'A{current_row}'].font = Font(bold=True)
    current_row += 1
    
    # Write pivot table for tag keys
    for col, column_name in enumerate(pivot_tag_keys.columns, start=2):
        tag_sheet.cell(row=current_row, column=col).value = column_name
        tag_sheet.cell(row=current_row, column=col).font = Font(bold=True)
    
    for row, (index_name, row_data) in enumerate(pivot_tag_keys.iterrows(), start=current_row+1):
        tag_sheet.cell(row=row, column=1).value = index_name
        for col, value in enumerate(row_data, start=2):
            cell = tag_sheet.cell(row=row, column=col)
            cell.value = value
            cell.number_format = '$#,##0.00'
    
    current_row += len(pivot_tag_keys) + 3
    
    # Write pivot tables for each tag key's values
    for tag_key, pivot_df in tag_value_pivots.items():
        tag_sheet[f'A{current_row}'] = f'Costs by {tag_key}'
        tag_sheet[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
        
        for col, column_name in enumerate(pivot_df.columns, start=2):
            tag_sheet.cell(row=current_row, column=col).value = column_name
            tag_sheet.cell(row=current_row, column=col).font = Font(bold=True)
        
        for row, (index_name, row_data) in enumerate(pivot_df.iterrows(), start=current_row+1):
            tag_sheet.cell(row=row, column=1).value = index_name
            for col, value in enumerate(row_data, start=2):
                cell = tag_sheet.cell(row=row, column=col)
                cell.value = value
                cell.number_format = '$#,##0.00'
        
        current_row += len(pivot_df) + 3
    
    # Auto-adjust column widths
    for column in tag_sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            adjusted_width = (max_length + 2)
            tag_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

def create_optimization_recommendations(wb, monthly_costs, credential):
    """Create sheet with cost optimization recommendations from Azure Advisor"""
    print("\nFetching cost optimization recommendations from Azure Advisor...")
    
    # Create a new sheet for recommendations
    rec_sheet = wb.create_sheet("Cost Optimization", 3)
    
    # Set up headers
    headers = ['Subscription', 'Resource', 'Impact', 'Problem', 'Recommendation', 'Potential Savings']
    for col, header in enumerate(headers, 1):
        cell = rec_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    row = 2
    subscription_names = {sub['id']: sub['name'] for sub in get_subscriptions()}
    
    # Get recommendations for each subscription
    for sub_id in monthly_costs[list(monthly_costs.keys())[0]].keys():
        print(f"\nFetching recommendations for subscription: {subscription_names.get(sub_id, sub_id)}")
        
        # Create Advisor client for this subscription
        advisor_client = AdvisorManagementClient(credential, sub_id)
        
        try:
            # Get cost recommendations
            recommendations = advisor_client.recommendations.list(
                filter="Category eq 'Cost'"
            )
            
            for recommendation in recommendations:
                # Extract potential savings if available
                try:
                    savings = float(recommendation.extended_properties.get('savingsAmount', 0))
                except (ValueError, TypeError):
                    savings = 0
                
                # Add recommendation to sheet
                rec_sheet.cell(row=row, column=1).value = subscription_names.get(sub_id, sub_id)
                rec_sheet.cell(row=row, column=2).value = recommendation.impacted_value
                rec_sheet.cell(row=row, column=3).value = recommendation.impact
                rec_sheet.cell(row=row, column=4).value = recommendation.short_description.problem
                rec_sheet.cell(row=row, column=5).value = recommendation.short_description.solution
                rec_sheet.cell(row=row, column=6).value = savings
                rec_sheet.cell(row=row, column=6).number_format = '$#,##0.00'
                
                # Add conditional formatting for impact
                if recommendation.impact == 'High':
                    rec_sheet.cell(row=row, column=3).fill = PatternFill(
                        start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
                    )
                elif recommendation.impact == 'Medium':
                    rec_sheet.cell(row=row, column=3).fill = PatternFill(
                        start_color="FFE4B5", end_color="FFE4B5", fill_type="solid"
                    )
                
                row += 1
                
        except Exception as e:
            print(f"Error getting recommendations for subscription {sub_id}: {str(e)}")
    
    # Auto-adjust column widths
    for column in rec_sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        rec_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

def create_excel_report(start_period=None, end_period=None):
    """
    Create Excel report with monthly sheets
    Args:
        start_period (dict): Dictionary containing 'year' and 'month' for start period
        end_period (dict): Dictionary containing 'year' and 'month' for end period
    Example:
        create_excel_report(
            start_period={'year': 2024, 'month': 11},
            end_period={'year': 2024, 'month': 12}
        )
    """
    if start_period is None or end_period is None:
        # Default to current month and previous month if no periods specified
        today = datetime.now()
        end_period = {'year': today.year, 'month': today.month}
        start_period = {
            'year': today.year if today.month > 1 else today.year - 1,
            'month': today.month - 1 if today.month > 1 else 12
        }
    
    print(f"Generating cost report from {calendar.month_name[start_period['month']]} {start_period['year']} "
          f"to {calendar.month_name[end_period['month']]} {end_period['year']}")
    
    # Create filename with start and end periods
    start_str = f"{start_period['year']}-{start_period['month']:02d}"
    end_str = f"{end_period['year']}-{end_period['month']:02d}"
    filename = f'azure_costs_{start_str}_to_{end_str}.xlsx'
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)
    
    # Store monthly costs for comparison
    monthly_costs = {}  # Format: {(year, month): {sub_id: {service_family: cost}}}
    
    # Create sheets for each month in the range
    start_date = datetime(start_period['year'], start_period['month'], 1)
    end_date = datetime(end_period['year'], end_period['month'], 1)
    
    current_date = start_date
    while current_date <= end_date:
        year = current_date.year
        month = current_date.month
        month_name = calendar.month_name[month]
        
        print(f"\nProcessing {month_name} {year}")
        sheet = wb.create_sheet(f"{month_name} {year}")
        
        # Set up headers
        sheet['A1'] = 'Subscription Name'
        sheet['A1'].font = Font(bold=True)
        
        # Get start and end dates for the month
        start_date_month = datetime(year, month, 1)
        end_date_month = start_date_month + relativedelta(months=1) - timedelta(days=1)
        
        # Format dates for Azure API
        date_format = '%Y-%m-%d'
        start_date_str = start_date_month.strftime(date_format)
        end_date_str = end_date_month.strftime(date_format)
        
        # Get unique resource types across all subscriptions for the month
        all_resource_types = set()
        for sub in get_subscriptions():
            costs = get_costs_by_resource_type(
                sub['id'], 
                start_date_str,
                end_date_str
            )
            all_resource_types.update(costs.keys())
        
        # Set up resource type columns
        for col, resource_type in enumerate(sorted(all_resource_types), start=2):
            cell = sheet.cell(row=1, column=col)
            cell.value = resource_type
            cell.font = Font(bold=True)
        
        # Add Total column header
        total_col = len(all_resource_types) + 2
        sheet.cell(row=1, column=total_col).value = 'Total'
        sheet.cell(row=1, column=total_col).font = Font(bold=True)
        
        # Initialize monthly costs for this month
        monthly_costs[(year, month)] = {}
        
        # Process each subscription
        for row, sub in enumerate(get_subscriptions(), start=2):
            sheet.cell(row=row, column=1).value = sub['name']
            
            costs = get_costs_by_resource_type(
                sub['id'],
                start_date_str,
                end_date_str
            )
            
            # Store costs for comparison
            monthly_costs[(year, month)][sub['id']] = costs
            
            # Initialize row total
            row_total = 0
            
            # Fill in costs for each resource type
            for col, resource_type in enumerate(sorted(all_resource_types), start=2):
                cell = sheet.cell(row=row, column=col)
                cost = costs.get(resource_type, 0)
                cell.value = cost
                cell.number_format = '$#,##0.00'
                row_total += cost
            
            # Add row total
            total_cell = sheet.cell(row=row, column=total_col)
            total_cell.value = row_total
            total_cell.number_format = '$#,##0.00'
            total_cell.font = Font(bold=True)
        
        # Add column totals at the bottom
        total_row = len(get_subscriptions()) + 2
        sheet.cell(row=total_row, column=1).value = 'Total'
        sheet.cell(row=total_row, column=1).font = Font(bold=True)
        
        # Calculate column totals
        for col in range(2, total_col + 1):
            column_total = sum(sheet.cell(row=r, column=col).value or 0 
                             for r in range(2, total_row))
            cell = sheet.cell(row=total_row, column=col)
            cell.value = column_total
            cell.number_format = '$#,##0.00'
            cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Move to next month
        current_date = current_date + relativedelta(months=1)
    
    # Create Summary sheet for cost changes
    summary_sheet = wb.create_sheet("Cost Changes", 0)
    
    # Set up summary headers
    summary_sheet['A1'] = 'Subscription'
    summary_sheet['B1'] = 'Service Family'
    summary_sheet['C1'] = 'Previous Month'
    summary_sheet['D1'] = 'Current Month'
    summary_sheet['E1'] = 'Change Amount'
    summary_sheet['F1'] = 'Change Percentage'
    summary_sheet['G1'] = 'Month Pair'
    
    # Apply header formatting
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
    
    # Calculate significant changes
    print("\nCalculating cost changes...")
    row = 2
    
    # Get sorted list of (year, month) tuples from monthly_costs
    periods = sorted(monthly_costs.keys())
    
    # Compare each consecutive pair of months
    for i in range(len(periods) - 1):
        prev_year, prev_month = periods[i]
        curr_year, curr_month = periods[i + 1]
        
        prev_month_name = calendar.month_name[prev_month]
        curr_month_name = calendar.month_name[curr_month]
        
        for sub in get_subscriptions():
            prev_costs = monthly_costs.get((prev_year, prev_month), {}).get(sub['id'], {})
            curr_costs = monthly_costs.get((curr_year, curr_month), {}).get(sub['id'], {})
            
            print(f"\nAnalyzing changes for {sub['name']}")
            print(f"Previous month ({prev_month_name} {prev_year}) costs: {prev_costs}")
            print(f"Current month ({curr_month_name} {curr_year}) costs: {curr_costs}")
            
            # Compare service families
            all_services = set(prev_costs.keys()) | set(curr_costs.keys())
            
            for service in all_services:
                prev_cost = prev_costs.get(service, 0)
                curr_cost = curr_costs.get(service, 0)
                
                # Calculate change
                change_amount = curr_cost - prev_cost
                if prev_cost > 0:
                    change_percentage = (change_amount / prev_cost) * 100
                else:
                    change_percentage = float('inf') if curr_cost > 0 else 0
                
                # Debug print
                print(f"Service: {service}")
                print(f"Previous cost: ${prev_cost:,.2f}")
                print(f"Current cost: ${curr_cost:,.2f}")
                print(f"Change: ${change_amount:,.2f} ({change_percentage:.1f}%)")
                
                # Only show significant changes (e.g., >10% and >$100)
                if abs(change_percentage) > 10 and abs(change_amount) > 100:
                    print(f"Significant change detected - adding to summary")
                    summary_sheet[f'A{row}'] = sub['name']
                    summary_sheet[f'B{row}'] = service
                    summary_sheet[f'C{row}'] = prev_cost
                    summary_sheet[f'D{row}'] = curr_cost
                    summary_sheet[f'E{row}'] = change_amount
                    summary_sheet[f'F{row}'] = f"{change_percentage:.1f}%"
                    summary_sheet[f'G{row}'] = f"{prev_month_name} {prev_year} → {curr_month_name} {curr_year}"
                    
                    # Apply conditional formatting
                    if change_amount > 0:
                        summary_sheet[f'E{row}'].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    else:
                        summary_sheet[f'E{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    
                    # Format numbers
                    for col in ['C', 'D', 'E']:
                        summary_sheet[f'{col}{row}'].number_format = '$#,##0.00'
                    
                    row += 1
    
    # Auto-adjust column widths for summary sheet
    for column in summary_sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        summary_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Add visualizations before saving
    create_cost_visualizations(wb, monthly_costs)
    
    # Add anomaly analysis after visualizations
    create_anomaly_analysis(wb, monthly_costs)
    
    # Get credential once
    credential = DefaultAzureCredential()
    
    # Add tag analysis after anomaly analysis
    create_tag_analysis(wb, monthly_costs, credential)
    
    # Add optimization recommendations with credential
    create_optimization_recommendations(wb, monthly_costs, credential)
    
    # Save workbook
    wb.save(filename)
    print(f"Report saved as {filename}")

if __name__ == "__main__":
    create_excel_report(
        start_period={'year': 2024, 'month': 12},  # December 2024
        end_period={'year': 2025, 'month': 1}      # February 2025
    ) 
    
